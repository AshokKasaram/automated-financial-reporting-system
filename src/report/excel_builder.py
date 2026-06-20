import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
import pandas as pd
from datetime import datetime

# ── Styles ────────────────────────────────────────────────────────────────────

NAVY_FILL   = PatternFill("solid", fgColor="1F3A5F")
BLUE_FILL   = PatternFill("solid", fgColor="2E6DA4")
ALT_FILL    = PatternFill("solid", fgColor="EBF2FA")
RED_FILL    = PatternFill("solid", fgColor="FDECEA")
GREEN_FILL  = PatternFill("solid", fgColor="E8F5E9")
AMBER_FILL  = PatternFill("solid", fgColor="FFF8E1")
GRAY_FILL   = PatternFill("solid", fgColor="F5F5F5")

WHITE_BOLD  = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
TITLE_FONT  = Font(color="1F3A5F", bold=True, size=14, name="Calibri")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
BODY_FONT   = Font(size=10, name="Calibri")
BOLD_BODY   = Font(bold=True, size=10, name="Calibri")

THIN  = Side(style="thin",   color="CCCCCC")
MED   = Side(style="medium", color="2E6DA4")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOTTOM_MED  = Border(bottom=MED)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")

MONEY_FMT   = '#,##0.00'
PCT_FMT     = '0.00"%"'
INT_FMT     = '#,##0'


def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header_row(ws, headers: list, row: int = 1):
    for col, title in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=title)
        c.fill      = NAVY_FILL
        c.font      = HEADER_FONT
        c.alignment = CENTER
        c.border    = BORDER


def _title_block(ws, title: str, subtitle: str, row: int = 1):
    ws.row_dimensions[row].height = 28
    ws.row_dimensions[row + 1].height = 18
    c = ws.cell(row=row, column=1, value=title)
    c.font = TITLE_FONT
    c.alignment = LEFT
    s = ws.cell(row=row + 1, column=1, value=subtitle)
    s.font = Font(color="666666", size=10, italic=True, name="Calibri")
    s.alignment = LEFT


# ── Sheet 1: P&L Summary ─────────────────────────────────────────────────────

def build_pl_sheet(ws, df: pd.DataFrame, month_label: str):
    ws.title = "P&L Summary"
    _set_col_widths(ws, [18, 24, 16, 16, 16, 14, 16])
    _title_block(ws, f"Profit & Loss Summary — {month_label}", "All figures in USD")
    ws.row_dimensions[3].height = 4  # spacer

    _header_row(ws, ["Category", "Account", "Budget", "Actual", "Variance", "Var %", "Status"], row=4)

    grouped = (
        df.groupby(["category", "account"])[["budget", "actual", "variance"]]
        .sum()
        .reset_index()
    )
    grouped["variance_pct"] = (grouped["variance"] / grouped["budget"] * 100).round(2)
    grouped["status"] = grouped["variance_pct"].apply(
        lambda v: "Over budget" if v > 10 else ("Under budget" if v < -10 else "On track")
    )

    status_fills = {
        "Over budget":   RED_FILL,
        "Under budget":  AMBER_FILL,
        "On track":      GREEN_FILL,
    }

    current_category = None
    for i, (_, row) in enumerate(grouped.iterrows(), start=5):
        alt = ALT_FILL if i % 2 == 0 else None

        # Category grouping label
        if row.category != current_category:
            current_category = row.category
            label_row = i
            c = ws.cell(row=label_row, column=1, value=row.category.upper())
            c.font      = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
            c.fill      = BLUE_FILL
            c.alignment = LEFT
            for col in range(2, 8):
                ws.cell(row=label_row, column=col).fill = BLUE_FILL
            i += 1

        values = [
            "",
            row.account,
            round(row.budget, 2),
            round(row.actual, 2),
            round(row.variance, 2),
            round(row.variance_pct, 2),
            row.status,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BORDER
            cell.font   = BODY_FONT
            if alt: cell.fill = alt
            if col == 2: cell.alignment = LEFT
            if col in (3, 4):
                cell.number_format = MONEY_FMT
                cell.alignment = RIGHT
            if col == 5:
                cell.number_format = MONEY_FMT
                cell.alignment = RIGHT
                cell.fill = RED_FILL if val < 0 else GREEN_FILL
                cell.font = Font(bold=True, size=10, name="Calibri",
                                 color="C62828" if val < 0 else "2E7D32")
            if col == 6:
                cell.number_format = PCT_FMT
                cell.alignment = RIGHT
            if col == 7:
                cell.fill = status_fills.get(val, alt or PatternFill())
                cell.alignment = CENTER
                cell.font = BOLD_BODY

    # Totals row
    last_row = ws.max_row + 1
    total_budget = grouped["budget"].sum()
    total_actual = grouped["actual"].sum()
    total_var    = total_actual - total_budget

    totals = ["", "TOTAL", total_budget, total_actual, total_var,
              round(total_var / total_budget * 100, 2), ""]
    for col, val in enumerate(totals, start=1):
        cell = ws.cell(row=last_row, column=col, value=val)
        cell.font   = BOLD_BODY
        cell.border = Border(top=MED, bottom=MED)
        cell.fill   = GRAY_FILL
        if col in (3, 4, 5): cell.number_format = MONEY_FMT; cell.alignment = RIGHT
        if col == 6:          cell.number_format = PCT_FMT;   cell.alignment = RIGHT


# ── Sheet 2: Budget vs Actuals ────────────────────────────────────────────────

def build_bva_sheet(ws, df: pd.DataFrame, month_label: str):
    ws.title = "Budget vs Actuals"
    _set_col_widths(ws, [10, 22, 16, 16, 16, 14])
    _title_block(ws, f"Budget vs Actuals — {month_label}", "Monthly breakdown by account")

    _header_row(ws, ["Month", "Account", "Budget", "Actual", "Variance", "Var %"], row=4)

    monthly = (
        df.groupby(["month", "month_name", "account"])[["budget", "actual", "variance"]]
        .sum()
        .reset_index()
        .sort_values(["month", "account"])
    )
    monthly["variance_pct"] = (monthly["variance"] / monthly["budget"] * 100).round(2)

    for i, (_, row) in enumerate(monthly.iterrows(), start=5):
        alt = ALT_FILL if i % 2 == 0 else None
        values = [
            row.month_name,
            row.account,
            round(row.budget, 2),
            round(row.actual, 2),
            round(row.variance, 2),
            round(row.variance_pct, 2),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BORDER
            cell.font   = BODY_FONT
            if alt: cell.fill = alt
            if col == 1: cell.alignment = CENTER
            if col == 2: cell.alignment = LEFT
            if col in (3, 4, 5):
                cell.number_format = MONEY_FMT
                cell.alignment = RIGHT
            if col == 5:
                cell.fill = RED_FILL if val < 0 else GREEN_FILL
                cell.font = Font(bold=True, size=10, name="Calibri",
                                 color="C62828" if val < 0 else "2E7D32")
            if col == 6:
                cell.number_format = PCT_FMT
                cell.alignment = RIGHT

    # Bar chart: total budget vs actual per month
    monthly_totals = (
        df.groupby(["month", "month_name"])[["budget", "actual"]]
        .sum()
        .reset_index()
        .sort_values("month")
    )
    chart_start_row = ws.max_row + 3
    ws.cell(row=chart_start_row,     column=1, value="Month")
    ws.cell(row=chart_start_row,     column=2, value="Budget")
    ws.cell(row=chart_start_row,     column=3, value="Actual")
    for j, (_, r) in enumerate(monthly_totals.iterrows(), start=chart_start_row + 1):
        ws.cell(row=j, column=1, value=r.month_name)
        ws.cell(row=j, column=2, value=round(r.budget, 2))
        ws.cell(row=j, column=3, value=round(r.actual, 2))

    chart = BarChart()
    chart.type    = "col"
    chart.title   = "Monthly Budget vs Actual"
    chart.style   = 10
    chart.y_axis.title = "Amount (USD)"
    chart.x_axis.title = "Month"

    data = Reference(ws, min_col=2, max_col=3,
                     min_row=chart_start_row,
                     max_row=chart_start_row + 12)
    cats = Reference(ws, min_col=1,
                     min_row=chart_start_row + 1,
                     max_row=chart_start_row + 12)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width  = 22
    chart.height = 14
    ws.add_chart(chart, "H4")


# ── Sheet 3: Variance Analysis ────────────────────────────────────────────────

def build_variance_sheet(ws, df: pd.DataFrame, month_label: str):
    ws.title = "Variance Analysis"
    _set_col_widths(ws, [18, 24, 14, 14, 20, 22])
    _title_block(ws, f"Variance Analysis — {month_label}", "Accounts sorted by absolute variance %")

    _header_row(ws, ["Category", "Account", "Avg Var %", "Total Var $", "Status", "Action needed"], row=4)

    summary = (
        df.groupby(["category", "account"])
        .agg(avg_var_pct=("variance_pct", "mean"), total_var=("variance", "sum"))
        .reset_index()
        .sort_values("avg_var_pct", key=abs, ascending=False)
    )
    summary["status"] = summary["avg_var_pct"].apply(
        lambda v: "Over budget" if v > 10 else ("Under budget" if v < -10 else "On track")
    )
    summary["action"] = summary["avg_var_pct"].apply(
        lambda v: "Review spend — exceeding plan"  if v > 10
             else ("Investigate shortfall" if v < -10 else "No action required")
    )

    status_fills = {
        "Over budget":  RED_FILL,
        "Under budget": AMBER_FILL,
        "On track":     GREEN_FILL,
    }

    for i, (_, row) in enumerate(summary.iterrows(), start=5):
        alt = ALT_FILL if i % 2 == 0 else None
        values = [
            row.category,
            row.account,
            round(row.avg_var_pct, 2),
            round(row.total_var, 2),
            row.status,
            row.action,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BORDER
            cell.font   = BODY_FONT
            if alt: cell.fill = alt
            if col in (1, 2, 6): cell.alignment = LEFT
            if col == 3:
                cell.number_format = PCT_FMT
                cell.alignment = RIGHT
            if col == 4:
                cell.number_format = MONEY_FMT
                cell.alignment = RIGHT
                cell.fill = RED_FILL if val < 0 else GREEN_FILL
                cell.font = Font(bold=True, size=10, name="Calibri",
                                 color="C62828" if val < 0 else "2E7D32")
            if col == 5:
                cell.fill      = status_fills.get(val, alt or PatternFill())
                cell.alignment = CENTER
                cell.font      = BOLD_BODY


# ── Main entry point ──────────────────────────────────────────────────────────

def generate_report(df: pd.DataFrame, output_path: str, month_label: str = None) -> str:
    if month_label is None:
        month_label = datetime.now().strftime("%B %Y")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_pl_sheet(wb.create_sheet(),       df, month_label)
    build_bva_sheet(wb.create_sheet(),      df, month_label)
    build_variance_sheet(wb.create_sheet(), df, month_label)

    wb.save(output_path)
    print(f"[report] Saved → {output_path}")
    return output_path
